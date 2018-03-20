# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import  load_workbook
import load_to_sqlite
import os


def readfile(filename, conn, tag_name):

    wb = load_workbook(filename)

    ws = wb.worksheets[0]
    data = {}
    for i in range(2, 200):
        data[i] = {}
        for j in range(1,7):
            value = ws.cell(i,j).value
            data[i][j] = value

    for i in range(2,200):
        '''print str(i-1)'''
        q = data[i][1]

        if q and q.strip():

            try:
                correct_index = int(data[i][2])
            except:
                correct_index = -1
                '''raw_input("ko co cau tra loi")'''

            a1 = data[i][3]
            a2 = data[i][4]
            a3= data[i][5]
            a4 = data[i][6]

            load_to_sqlite.insert_to_db(conn,question=q, a1=a1,a2=a2,a3=a3,a4=a4,correct_answer=correct_index, tag_name=tag_name)




        else:

            print tag_name + ' :' + str(i-1)
            break

if __name__ == '__main__':
    all_dir = r'C:\Users\vuth1\OneDrive\Documents\On tap Thi Nang luc Ha Tinh 2018 (Moi)\On tap Thi Nang luc Ha Tinh 2018 (Moi)\Vien thong\all'

    conn = load_to_sqlite.open_connection()
    for filename in os.listdir(all_dir):
        tag_name = filename.split(r'.')[0].strip()
        readfile(all_dir + r'\\' + filename, conn, tag_name=tag_name)

    conn.commit()
    conn.close()